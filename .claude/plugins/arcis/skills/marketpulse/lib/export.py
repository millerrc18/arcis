"""Export module for MarketPulse analytics results.

Provides stateless functions for writing analytics output to Excel, CSV,
Parquet, and JSON formats.  Each public function accepts either a
``pd.DataFrame`` or an :class:`AnalyticsResult` instance.

Public API
----------
- :func:`to_excel`   -- formatted multi-sheet workbook
- :func:`to_csv`     -- single CSV file
- :func:`to_parquet` -- Parquet file
- :func:`to_json`    -- JSON-serializable dict (no file written)
"""

from __future__ import annotations

import dataclasses
import math
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .analytics._base import AnalyticsResult

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(color="006100")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT = Font(color="9C0006")

_PCT_FORMAT = "0.00%"
_NUM_FORMAT = "#,##0.00"
_INT_FORMAT = "#,##0"

# Column name substrings that map to specific number formats
_PCT_KEYWORDS = ("return", "pct", "volatil", "correlation")
_INT_KEYWORDS = ("volume", "num_transactions", "bar_count", "sample_count", "ticker_count",
                 "date_count", "num_periods", "spike_ratio")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _auto_column_widths(ws: "Worksheet", min_width: int = 10, max_width: int = 40) -> None:
    """Set each column width based on the maximum content length in that column."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = adjusted


def _style_header_row(ws: "Worksheet") -> None:
    """Apply bold white text on blue fill to the first row; freeze pane at A2."""
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _apply_return_formatting(
    ws: "Worksheet", col_idx: int, start_row: int = 2
) -> None:
    """Green/red conditional formatting for return/percentage columns.

    Positive values receive green fill + dark-green font; negative values
    receive red fill + dark-red font.  ``col_idx`` is 1-based.
    """
    for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value is None:
                continue
            try:
                val = float(cell.value)
            except (TypeError, ValueError):
                continue
            if val > 0:
                cell.fill = _GREEN_FILL
                cell.font = _GREEN_FONT
            elif val < 0:
                cell.fill = _RED_FILL
                cell.font = _RED_FONT


def _apply_number_formats(ws: "Worksheet", df: pd.DataFrame) -> None:
    """Infer and apply number formats to data cells based on column names.

    Rules (applied in order, first match wins per column):
    - Column name contains a PCT keyword → percentage format ``0.00%``
    - Column name contains an INT keyword → integer-with-commas ``#,##0``
    - Column dtype is float → ``#,##0.00``

    Only data rows (row 2 onward) are affected; the header row is skipped.
    """
    for col_idx, col_name in enumerate(df.columns, start=1):
        lower = col_name.lower()

        if any(kw in lower for kw in _PCT_KEYWORDS):
            fmt = _PCT_FORMAT
        elif any(kw in lower for kw in _INT_KEYWORDS):
            fmt = _INT_FORMAT
        elif pd.api.types.is_float_dtype(df[col_name]):
            fmt = _NUM_FORMAT
        else:
            continue  # strings, ints, etc. — no special format needed

        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt


def _result_to_dataframe(result: AnalyticsResult) -> pd.DataFrame:
    """Convert an AnalyticsResult to a flat DataFrame.

    Strategy:
    1. Find the largest field whose value is a non-empty list of dicts.
       Expand those dicts into rows; attach any scalar fields as extra columns.
    2. If no list-of-dicts field exists, build a single-row DataFrame from
       all scalar fields (flattening nested dicts one level).
    """
    if not dataclasses.is_dataclass(result):
        raise TypeError(f"Expected a dataclass AnalyticsResult, got {type(result)}")

    fields = dataclasses.fields(result)  # type: ignore[arg-type]
    scalar_fields: dict = {}
    best_list_field: str | None = None
    best_list_val: list | None = None

    for field in fields:
        val = getattr(result, field.name)
        if (
            isinstance(val, list)
            and val
            and isinstance(val[0], (dict, AnalyticsResult))
        ):
            if best_list_val is None or len(val) > len(best_list_val):
                best_list_field = field.name
                best_list_val = val
        else:
            # Scalar field (str, int, float, list of non-dicts, etc.)
            if not isinstance(val, list):
                scalar_fields[field.name] = val

    if best_list_val is not None:
        rows = []
        for item in best_list_val:
            if dataclasses.is_dataclass(item):
                row = dataclasses.asdict(item)  # type: ignore[arg-type]
            else:
                row = dict(item)
            rows.append(row)
        df = pd.DataFrame(rows)
        # Attach scalar fields as constant columns (skip the list-field name)
        for key, sval in scalar_fields.items():
            if key != best_list_field:
                df[key] = sval
        return df

    # No expandable list -- single-row from all scalar fields
    row: dict = {}
    for field in fields:
        val = getattr(result, field.name)
        if isinstance(val, dict):
            for k, v in val.items():
                row[f"{field.name}.{k}"] = v
        elif isinstance(val, list):
            row[field.name] = str(val)
        else:
            row[field.name] = val
    return pd.DataFrame([row])


def _default_output_path(basename: str, ext: str) -> Path:
    """Build a timestamped output path on the user's Desktop (fallback: home)."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{basename}_{ts}.{ext}"
    desktop = Path.home() / "Desktop"
    if not desktop.exists():
        desktop = Path.home()
    return desktop / filename


def _write_df_to_sheet(
    ws: "Worksheet",
    df: pd.DataFrame,
    return_columns: list[str] | None = None,
) -> None:
    """Write a DataFrame to an openpyxl worksheet with full formatting.

    Steps
    -----
    1. Write header row.
    2. Write data rows, converting numpy/pandas scalar types to Python native
       and mapping NaN → None.
    3. Apply header styling, number formats, and (optionally) return coloring.
    4. Auto-size columns.
    """
    # --- Header ---
    ws.append(list(df.columns))

    # --- Data rows ---
    for _, row in df.iterrows():
        out_row = []
        for val in row:
            if val is None:
                out_row.append(None)
            elif hasattr(val, "item"):
                # numpy scalar → Python native
                try:
                    native = val.item()
                except Exception:
                    native = val
                out_row.append(native)
            elif isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                out_row.append(None)
            else:
                try:
                    if pd.isna(val):
                        out_row.append(None)
                        continue
                except (TypeError, ValueError):
                    pass
                out_row.append(val)
        ws.append(out_row)

    # --- Styling ---
    _style_header_row(ws)
    _apply_number_formats(ws, df)

    if return_columns:
        for col_name in return_columns:
            if col_name in df.columns:
                col_idx = list(df.columns).index(col_name) + 1
                _apply_return_formatting(ws, col_idx)

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def to_excel(
    data: pd.DataFrame | AnalyticsResult,
    path: str | Path | None = None,
    sheets: dict[str, pd.DataFrame] | None = None,
    sheet_name: str = "Data",
    return_columns: list[str] | None = None,
) -> Path:
    """Write data to a formatted Excel workbook and return the output path.

    Parameters
    ----------
    data:
        Primary DataFrame or AnalyticsResult to write.  If *sheets* is also
        provided, *data* is written to a sheet named *sheet_name* in addition
        to any sheets in *sheets*.
    path:
        Output file path.  When ``None``, a timestamped file is created on the
        user's Desktop.
    sheets:
        Additional ``{sheet_name: DataFrame}`` mapping for multi-sheet output.
        These are written before the primary *data* sheet.
    sheet_name:
        Name for the worksheet that receives *data*.  Defaults to ``"Data"``.
    return_columns:
        Column names that should receive green/red conditional formatting.

    Returns
    -------
    Path
        Resolved absolute path of the written file.
    """
    if isinstance(data, AnalyticsResult):
        df = _result_to_dataframe(data)
    else:
        df = data

    out = Path(path) if path is not None else _default_output_path("marketpulse_export", "xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default empty sheet
    if wb.active is not None:
        wb.remove(wb.active)

    # Write extra sheets first
    if sheets:
        for sname, sdf in sheets.items():
            ws = wb.create_sheet(title=sname)
            _write_df_to_sheet(ws, sdf, return_columns=return_columns)

    # Write primary data sheet
    ws_main = wb.create_sheet(title=sheet_name)
    _write_df_to_sheet(ws_main, df, return_columns=return_columns)

    wb.save(str(out))
    return out.resolve()


def to_csv(
    data: pd.DataFrame | AnalyticsResult,
    path: str | Path | None = None,
) -> Path:
    """Write data to a CSV file and return the output path.

    Parameters
    ----------
    data:
        DataFrame or AnalyticsResult to export.
    path:
        Output file path.  When ``None``, a timestamped file is created on the
        user's Desktop.

    Returns
    -------
    Path
        Resolved absolute path of the written file.
    """
    if isinstance(data, AnalyticsResult):
        df = _result_to_dataframe(data)
    else:
        df = data

    out = Path(path) if path is not None else _default_output_path("marketpulse_export", "csv")
    out.parent.mkdir(parents=True, exist_ok=True)

    df.to_csv(str(out), index=False)
    return out.resolve()


def to_parquet(
    data: pd.DataFrame | AnalyticsResult,
    path: str | Path | None = None,
) -> Path:
    """Write data to a Parquet file and return the output path.

    Parameters
    ----------
    data:
        DataFrame or AnalyticsResult to export.
    path:
        Output file path.  When ``None``, a timestamped file is created on the
        user's Desktop.

    Returns
    -------
    Path
        Resolved absolute path of the written file.
    """
    if isinstance(data, AnalyticsResult):
        df = _result_to_dataframe(data)
    else:
        df = data

    out = Path(path) if path is not None else _default_output_path("marketpulse_export", "parquet")
    out.parent.mkdir(parents=True, exist_ok=True)

    table = pa.Table.from_pandas(df, preserve_index=False)
    pq.write_table(table, str(out))
    return out.resolve()


def to_json(
    data: pd.DataFrame | AnalyticsResult,
) -> dict:
    """Convert data to a JSON-serializable dict.  No file is written.

    For :class:`AnalyticsResult`, delegates to :meth:`~AnalyticsResult.to_dict`
    which handles nested dataclass serialization and NaN/date handling.

    For a ``pd.DataFrame``, returns::

        {"rows": [...], "row_count": N}

    with ISO-format timestamps, ``None`` in place of NaN/Inf, and Python
    native types (no numpy scalars).

    Parameters
    ----------
    data:
        DataFrame or AnalyticsResult to serialize.

    Returns
    -------
    dict
        JSON-serializable dictionary.
    """
    if isinstance(data, AnalyticsResult):
        return data.to_dict()

    # DataFrame path
    records = []
    for _, row in data.iterrows():
        out_row: dict = {}
        for col, val in row.items():
            if hasattr(val, "item"):
                try:
                    val = val.item()
                except Exception:
                    pass
            if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                val = None
            else:
                try:
                    if pd.isna(val):
                        val = None
                except (TypeError, ValueError):
                    pass
            # Timestamps → isoformat
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            out_row[col] = val
        records.append(out_row)

    return {"rows": records, "row_count": len(records)}
