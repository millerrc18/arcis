"""Tests for MarketPulse export module (lib/export.py)."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
import pytest

# ---------------------------------------------------------------------------
# Ensure lib is importable
# ---------------------------------------------------------------------------
_MP_ROOT = Path(__file__).resolve().parent.parent
if str(_MP_ROOT) not in sys.path:
    sys.path.insert(0, str(_MP_ROOT))

from tests.fixtures.make_bars import make_bars_df  # noqa: E402
from lib.analytics.summary import daily_summary, volume_analysis  # noqa: E402
from lib.export import (  # noqa: E402
    to_excel,
    to_csv,
    to_parquet,
    to_json,
    _result_to_dataframe,
)


# ---------------------------------------------------------------------------
# Shared test data helpers
# ---------------------------------------------------------------------------

def _make_simple_df() -> pd.DataFrame:
    """Small 3-row DataFrame with a mix of column types."""
    return pd.DataFrame(
        {
            "ticker": ["AAPL", "MSFT", "GOOG"],
            "close": [150.25, 310.50, 2750.00],
            "daily_return": [0.012, -0.005, 0.031],
            "volume": [1_000_000, 800_000, 500_000],
        }
    )


def _make_daily_summary_result(tickers: list[str] = None, days: int = 3):
    """Build a DailySummaryResult with the given tickers and day count."""
    if tickers is None:
        tickers = ["AAPL", "MSFT"]
    df = make_bars_df(tickers=tickers, days=days, bars_per_day=10, seed=42)
    return daily_summary(df)


def _make_volume_result(tickers: list[str] = None):
    """Build a VolumeAnalysisResult with a single ticker."""
    if tickers is None:
        tickers = ["AAPL"]
    df = make_bars_df(tickers=tickers, days=2, bars_per_day=10, seed=77)
    return volume_analysis(df)


# ---------------------------------------------------------------------------
# TestToExcel
# ---------------------------------------------------------------------------


class TestToExcel:
    """Tests for ``to_excel()``."""

    def test_excel_from_dataframe(self, tmp_path: Path):
        """Export a DataFrame and verify the file exists with correct dimensions."""
        import openpyxl

        df = _make_simple_df()
        out = to_excel(df, path=tmp_path / "out.xlsx")

        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # 1 header row + 3 data rows
        assert ws.max_row == 4
        # 4 columns in the DataFrame
        assert ws.max_column == len(df.columns)

    def test_excel_from_analytics_result(self, tmp_path: Path):
        """Export a DailySummaryResult, verify row count = 2 tickers * 3 days + 1 header."""
        import openpyxl

        result = _make_daily_summary_result(tickers=["AAPL", "MSFT"], days=3)
        out = to_excel(result, path=tmp_path / "summary.xlsx")

        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # 2 tickers * 3 days = 6 data rows + 1 header row = 7 total rows
        assert ws.max_row == 7

    def test_excel_multi_sheet(self, tmp_path: Path):
        """Multi-sheet export includes both extra sheet and primary data sheet."""
        import openpyxl

        df = _make_simple_df()
        summary_df = pd.DataFrame({"metric": ["count"], "value": [3]})
        out = to_excel(
            df,
            path=tmp_path / "multi.xlsx",
            sheets={"Summary": summary_df},
            sheet_name="Data",
        )

        wb = openpyxl.load_workbook(str(out))
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Data" in sheet_names

    def test_excel_header_formatting(self, tmp_path: Path):
        """Header cells should be bold with blue fill (#2F5496)."""
        import openpyxl

        df = _make_simple_df()
        out = to_excel(df, path=tmp_path / "fmt.xlsx")

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        for cell in ws[1]:
            assert cell.font.bold, f"Header cell {cell.coordinate} is not bold"
            # Fill color should be 2F5496 (the blue set in export.py)
            assert cell.fill.fgColor.rgb.upper().endswith("2F5496"), (
                f"Header cell {cell.coordinate} fill color is {cell.fill.fgColor.rgb}"
            )

    def test_excel_frozen_panes(self, tmp_path: Path):
        """Freeze pane should be set to A2 (freeze the header row)."""
        import openpyxl

        df = _make_simple_df()
        out = to_excel(df, path=tmp_path / "frozen.xlsx")

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_excel_return_column_formatting(self, tmp_path: Path):
        """Return columns should have green or red fills on data cells."""
        import openpyxl

        df = _make_simple_df()
        # daily_return has values 0.012, -0.005, 0.031
        out = to_excel(df, path=tmp_path / "retfmt.xlsx", return_columns=["daily_return"])

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active

        # Find the column index for daily_return (1-based)
        col_idx = list(df.columns).index("daily_return") + 1

        colored_cells = 0
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            fill_rgb = cell.fill.fgColor.rgb.upper()
            # openpyxl may emit ARGB with alpha prefix "00" or "FF"; check
            # the last 6 hex digits to be alpha-independent.
            color_hex = fill_rgb[-6:]
            if color_hex in ("C6EFCE", "FFC7CE"):
                colored_cells += 1

        # At least one cell should have a color fill
        assert colored_cells >= 1, (
            "Expected at least one green/red fill in daily_return column"
        )

    def test_excel_default_path(self):
        """When path=None, a timestamped file is created on the Desktop."""
        df = _make_simple_df()
        out = to_excel(df, path=None)

        try:
            assert out.exists()
            assert out.name.startswith("marketpulse_")
            assert out.suffix == ".xlsx"
            # Should be on the Desktop (or home if Desktop doesn't exist)
            desktop = Path.home() / "Desktop"
            expected_parent = desktop if desktop.exists() else Path.home()
            assert out.parent == expected_parent
        finally:
            if out.exists():
                out.unlink()


# ---------------------------------------------------------------------------
# TestToCsv
# ---------------------------------------------------------------------------


class TestToCsv:
    """Tests for ``to_csv()``."""

    def test_csv_from_dataframe(self, tmp_path: Path):
        """Export DataFrame, read back, verify row count and column names."""
        df = _make_simple_df()
        out = to_csv(df, path=tmp_path / "data.csv")

        assert out.exists()
        loaded = pd.read_csv(str(out))
        assert len(loaded) == len(df)
        assert list(loaded.columns) == list(df.columns)

    def test_csv_from_analytics_result(self, tmp_path: Path):
        """Export VolumeAnalysisResult (1 ticker) and verify row count = 1."""
        result = _make_volume_result(tickers=["AAPL"])
        out = to_csv(result, path=tmp_path / "volume.csv")

        assert out.exists()
        loaded = pd.read_csv(str(out))
        # 1 ticker -> 1 VolumeStats row
        assert len(loaded) == 1


# ---------------------------------------------------------------------------
# TestToParquet
# ---------------------------------------------------------------------------


class TestToParquet:
    """Tests for ``to_parquet()``."""

    def test_parquet_from_dataframe(self, tmp_path: Path):
        """Export DataFrame, read back, verify row count."""
        df = _make_simple_df()
        out = to_parquet(df, path=tmp_path / "data.parquet")

        assert out.exists()
        loaded = pd.read_parquet(str(out))
        assert len(loaded) == len(df)

    def test_parquet_from_analytics_result(self, tmp_path: Path):
        """Export DailySummaryResult (2 tickers * 3 days = 6 rows), verify row count."""
        result = _make_daily_summary_result(tickers=["AAPL", "MSFT"], days=3)
        out = to_parquet(result, path=tmp_path / "summary.parquet")

        assert out.exists()
        loaded = pd.read_parquet(str(out))
        assert len(loaded) == 6


# ---------------------------------------------------------------------------
# TestToJson
# ---------------------------------------------------------------------------


class TestToJson:
    """Tests for ``to_json()``."""

    def test_json_from_analytics_result(self):
        """DailySummaryResult.to_json() should return a dict with 'summaries' key."""
        result = _make_daily_summary_result(tickers=["AAPL", "MSFT"], days=3)
        d = to_json(result)

        assert isinstance(d, dict)
        assert "summaries" in d

    def test_json_from_dataframe(self):
        """DataFrame path should return dict with 'rows' and 'row_count' keys."""
        df = _make_simple_df()
        d = to_json(df)

        assert isinstance(d, dict)
        assert "rows" in d
        assert "row_count" in d
        assert d["row_count"] == len(df)
        assert isinstance(d["rows"], list)
        assert len(d["rows"]) == len(df)

    def test_json_serializable(self):
        """Both DataFrame and AnalyticsResult output should round-trip through json.dumps."""
        # DataFrame path
        df = _make_simple_df()
        d_df = to_json(df)
        serialized = json.dumps(d_df)
        assert isinstance(serialized, str)

        # AnalyticsResult path
        result = _make_daily_summary_result(tickers=["AAPL"], days=2)
        d_result = to_json(result)
        serialized2 = json.dumps(d_result)
        assert isinstance(serialized2, str)


# ---------------------------------------------------------------------------
# TestResultToDataframe
# ---------------------------------------------------------------------------


class TestResultToDataframe:
    """Tests for ``_result_to_dataframe()`` internal helper."""

    def test_list_field_expansion(self):
        """DailySummaryResult with 2 tickers * 3 days expands to 6 rows."""
        result = _make_daily_summary_result(tickers=["AAPL", "MSFT"], days=3)
        df = _result_to_dataframe(result)

        # 2 tickers * 3 days = 6 rows
        assert len(df) == 6
        # Should have columns from DailySummary fields
        assert "ticker" in df.columns
        assert "date" in df.columns
        assert "daily_return" in df.columns

    def test_scalar_result_single_row(self):
        """A result with no list-of-dataclass field becomes a single-row DataFrame."""
        from lib.analytics.types import GarmanKlassResult

        gk = GarmanKlassResult(
            ticker="AAPL",
            gk_vol=0.015,
            annualized_gk_vol=0.238,
            num_bars=390,
        )
        df = _result_to_dataframe(gk)

        assert len(df) == 1
        assert "ticker" in df.columns
        assert df["ticker"].iloc[0] == "AAPL"

    def test_raises_on_non_dataclass(self):
        """Passing a plain dict should raise TypeError."""
        with pytest.raises(TypeError):
            _result_to_dataframe({"key": "value"})  # type: ignore[arg-type]

    def test_volume_analysis_expands_stats(self):
        """VolumeAnalysisResult (1 ticker) should produce a 1-row DataFrame."""
        result = _make_volume_result(tickers=["AAPL"])
        df = _result_to_dataframe(result)

        assert len(df) == 1
        assert "ticker" in df.columns
        assert df["ticker"].iloc[0] == "AAPL"
